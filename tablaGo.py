from openpyxl import load_workbook
import json
from io import BytesIO

class tablaGo:
    def __init__(self, input_excel_file, prompts):
        #Path al archivo .xlsx
        self.input_excel_file = load_workbook(input_excel_file)

        #Dirección al archivo json con la info
        with open(prompts, 'r', encoding='utf-8') as f:
            self.prompts = json.load(f)

        self.questions = self.prompts["B1"]["preguntas"]
        self.casillas = self.prompts["B1"]["casillas"]

        self.file = load_workbook(input_excel_file)
        self.A1 = self.file["A1 Resumen"]
        self.B1 = self.file["B1 Requisitos licitación"]

        self.err = self.prompts["err_404"]
    # Rellena las casillas de la Ficha GO
    def update_excel(self, sheet, cell, answer):
        self.input_excel_file[sheet][cell] = answer


    def modify(self, sheet, cell, txt):
        self.file[sheet][cell] = txt

    def save_file(self):
        output = BytesIO()
        self.file.save(output)
        output.seek(0)  # Rewind the buffer
        return output

    def contains_any_phrases(self, input_string, phrases_list):
        for phrase in phrases_list:
            #print(f"{input_string} | {phrase}")
            if phrase in input_string:
                return True
        return False



def main():    
    input_xlsx = "ficha.xlsx"
    prompts = "prompts.json"
    tabla = tablaGo(input_xlsx, prompts)
    tabla.update_excel("PRUEBA", 'A1', "Hola")
    tabla.save("res.xlsx")

if __name__ == "__main__":
    main()