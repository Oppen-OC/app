from openpyxl import Workbook, load_workbook
from io import BytesIO
import streamlit as st


class tablaGo:
    def __init__(self, input_excel_file, cells, prompts):
        self.input_excel_file = input_excel_file
        self.cells = cells
        self.prompts = self.read_file_to_array(prompts)

    # Rellena las casillas de la Ficha GO
    def update_excel(self, input_excel_file, sheet_name, answer):
        # Load an existing Workbook
        wb = load_workbook(input_excel_file)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(ws)
        else:
            ws = wb.create_sheet(sheet_name)

        # Inserta texto en las casillas importantes
        for i in self.cells:
            cellB = 'B' + str(i)
            ws[cellB] = answer
        
        # Save the workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        #wb.save(input_excel_file[:-5] +"Res.xlsx")
        
    def read_file_to_array(self, file_path):
        """Reads a .txt file and returns its contents as a list of lines."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.readlines()
                # Strip newline characters and return as a list
                return [line.strip() for line in content]
        except FileNotFoundError:
            print(f"The file at {file_path} was not found.")
            return []
        except Exception as e:
            print(f"An error occurred: {e}")
            return

    def generate_responses(self):
        print("hello")
    
'''
def main():    
    input_xlsx = "ficha.xlsx"
    cells = [3, 4, 5, 6, 9, 10, 12, 13, 15, 17, 19, 21, 23, 24, 26, 27, 29, 31, 33, 35]
    prompts = "prompts.txt"
    tablaGo(input_xlsx, cells, prompts)
    return tablaGo.update_excel()

'''